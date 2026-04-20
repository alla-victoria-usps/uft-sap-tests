import csv
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


class ExcelManagerFrame(ttk.Frame):
    def __init__(self, parent, on_change=None):
        super().__init__(parent, style="Content.TFrame")
        self.on_change = on_change
        self.workbooks = {}
        self.file_paths = []
        self.current_file = None
        self.current_sheet = None
        self.row_index_map = []
        self._build_ui()

    def _build_ui(self):
        ttk.Label(self, text="Excel Document Management", style="Title.TLabel").pack(anchor="w", padx=16, pady=(16, 8))

        main = ttk.Frame(self, style="Content.TFrame")
        main.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        main.columnconfigure(1, weight=1)
        main.rowconfigure(1, weight=1)

        file_panel = ttk.LabelFrame(main, text="Files", style="Card.TLabelframe")
        file_panel.grid(row=0, column=0, rowspan=3, sticky="ns", padx=(0, 12))
        self.file_listbox = tk.Listbox(file_panel, height=16, exportselection=False)
        self.file_listbox.pack(fill="both", expand=True, padx=8, pady=8)
        self.file_listbox.bind("<<ListboxSelect>>", self._on_file_select)

        controls = ttk.Frame(main, style="Content.TFrame")
        controls.grid(row=0, column=1, sticky="ew", pady=(0, 8))
        controls.columnconfigure(12, weight=1)

        for idx, (label, command) in enumerate(
            [
                ("Import Excel", self.import_excel),
                ("Create New", self.create_new),
                ("Save", self.save_current),
                ("Save As", self.save_as),
                ("Export CSV", self.export_csv),
                ("Add Row", self.add_row),
                ("Delete Row", self.delete_row),
                ("Add Column", self.add_column),
                ("Delete Column", self.delete_column),
                ("Refresh", self.refresh),
            ]
        ):
            ttk.Button(controls, text=label, command=command).grid(row=0, column=idx, padx=(0, 6), sticky="w")

        sheet_row = ttk.Frame(main, style="Content.TFrame")
        sheet_row.grid(row=1, column=1, sticky="new")
        ttk.Label(sheet_row, text="Sheet", style="Field.TLabel").pack(side="left", padx=(0, 8))
        self.sheet_var = tk.StringVar()
        self.sheet_selector = ttk.Combobox(sheet_row, textvariable=self.sheet_var, state="readonly", width=30)
        self.sheet_selector.pack(side="left", padx=(0, 12))
        self.sheet_selector.bind("<<ComboboxSelected>>", self._on_sheet_select)

        ttk.Label(sheet_row, text="Search/Filter", style="Field.TLabel").pack(side="left", padx=(0, 8))
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(sheet_row, textvariable=self.search_var)
        search_entry.pack(side="left", fill="x", expand=True)
        search_entry.bind("<KeyRelease>", lambda _e: self._render_sheet())

        table_frame = ttk.Frame(main, style="Content.TFrame")
        table_frame.grid(row=2, column=1, sticky="nsew", pady=(8, 0))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(table_frame, show="headings")
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<Double-1>", self.edit_cell)

        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

    def _get_workbook(self):
        if not self.current_file:
            return None
        return self.workbooks.get(self.current_file)

    def _get_sheet(self):
        workbook = self._get_workbook()
        if not workbook:
            return None
        if self.current_sheet and self.current_sheet in workbook.sheetnames:
            return workbook[self.current_sheet]
        return workbook.active

    def _refresh_file_list(self):
        selected = self.current_file
        self.file_listbox.delete(0, tk.END)
        self.file_paths = list(self.workbooks.keys())
        for path in self.file_paths:
            self.file_listbox.insert(tk.END, Path(path).name)
        if selected and selected in self.file_paths:
            index = self.file_paths.index(selected)
            self.file_listbox.selection_clear(0, tk.END)
            self.file_listbox.selection_set(index)

    def _set_current_file(self, path: str):
        self.current_file = path
        workbook = self._get_workbook()
        if not workbook:
            return
        self.sheet_selector["values"] = workbook.sheetnames
        self.current_sheet = workbook.active.title
        self.sheet_var.set(self.current_sheet)
        self._render_sheet()
        if self.on_change:
            self.on_change()

    def _render_sheet(self):
        worksheet = self._get_sheet()
        if worksheet is None:
            self.tree["columns"] = ()
            return

        max_rows = max(worksheet.max_row, 1)
        max_cols = max(worksheet.max_column, 1)
        columns = [f"c{idx}" for idx in range(1, max_cols + 1)]
        self.tree["columns"] = columns
        for idx, col_id in enumerate(columns, start=1):
            label = get_column_letter(idx)
            self.tree.heading(col_id, text=label)
            self.tree.column(col_id, width=120, anchor="w")

        for item in self.tree.get_children():
            self.tree.delete(item)

        keyword = self.search_var.get().strip().lower()
        self.row_index_map = []
        for row_num in range(1, max_rows + 1):
            values = [worksheet.cell(row=row_num, column=col_idx).value for col_idx in range(1, max_cols + 1)]
            display_values = ["" if value is None else str(value) for value in values]
            if keyword and not any(keyword in cell.lower() for cell in display_values):
                continue
            self.row_index_map.append(row_num)
            self.tree.insert("", "end", values=display_values)

    def import_excel(self):
        files = filedialog.askopenfilenames(title="Import Excel Files", filetypes=[("Excel Files", "*.xlsx *.xls")])
        if not files:
            return
        imported = 0
        for file_path in files:
            resolved = str(Path(file_path).resolve())
            try:
                self.workbooks[resolved] = load_workbook(resolved)
                imported += 1
            except Exception as exc:
                messagebox.showerror("Import Failed", f"Could not import:\n{resolved}\n\n{exc}")
        if imported:
            self._refresh_file_list()
            self._set_current_file(next(reversed(self.workbooks)))
            messagebox.showinfo("Imported", f"Imported {imported} Excel file(s).")

    def create_new(self):
        sheet_name = simpledialog.askstring("New Workbook", "Enter initial sheet name:", initialvalue="Sheet1")
        if not sheet_name:
            return
        save_path = filedialog.asksaveasfilename(
            title="Create Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not save_path:
            return
        workbook = Workbook()
        workbook.active.title = sheet_name.strip() or "Sheet1"
        workbook.save(save_path)
        resolved = str(Path(save_path).resolve())
        self.workbooks[resolved] = workbook
        self._refresh_file_list()
        self._set_current_file(resolved)
        messagebox.showinfo("Created", f"Created workbook:\n{resolved}")

    def save_current(self):
        workbook = self._get_workbook()
        if workbook is None or not self.current_file:
            messagebox.showwarning("No File", "Import or create an Excel file first.")
            return
        workbook.save(self.current_file)
        messagebox.showinfo("Saved", f"Saved workbook:\n{self.current_file}")

    def save_as(self):
        workbook = self._get_workbook()
        if workbook is None:
            messagebox.showwarning("No File", "Import or create an Excel file first.")
            return
        save_path = filedialog.asksaveasfilename(
            title="Save Workbook As",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not save_path:
            return
        resolved = str(Path(save_path).resolve())
        workbook.save(resolved)
        self.workbooks[resolved] = workbook
        self._refresh_file_list()
        self._set_current_file(resolved)
        messagebox.showinfo("Saved", f"Workbook saved as:\n{resolved}")

    def export_csv(self):
        worksheet = self._get_sheet()
        if worksheet is None:
            messagebox.showwarning("No Sheet", "Select a workbook and sheet first.")
            return
        output = filedialog.asksaveasfilename(title="Export CSV", defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not output:
            return
        max_rows = max(worksheet.max_row, 1)
        max_cols = max(worksheet.max_column, 1)
        with open(output, "w", newline="", encoding="utf-8") as csv_file:
            writer = csv.writer(csv_file)
            for row_num in range(1, max_rows + 1):
                row_values = [worksheet.cell(row=row_num, column=col_idx).value for col_idx in range(1, max_cols + 1)]
                writer.writerow(["" if value is None else value for value in row_values])
        messagebox.showinfo("Exported", f"Exported sheet to CSV:\n{output}")

    def add_row(self):
        worksheet = self._get_sheet()
        if worksheet is None:
            messagebox.showwarning("No Sheet", "Select a workbook and sheet first.")
            return
        worksheet.append([None] * max(worksheet.max_column, 1))
        self._render_sheet()

    def delete_row(self):
        worksheet = self._get_sheet()
        if worksheet is None:
            messagebox.showwarning("No Sheet", "Select a workbook and sheet first.")
            return
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Select one or more rows to delete.")
            return
        indexes = sorted((self.tree.index(item) for item in selected), reverse=True)
        for index in indexes:
            if index < len(self.row_index_map):
                worksheet.delete_rows(self.row_index_map[index], 1)
        self._render_sheet()

    def _parse_column_index(self, value: str):
        candidate = value.strip()
        if not candidate:
            return None
        if candidate.isdigit():
            index = int(candidate)
            return index if index >= 1 else None
        try:
            return column_index_from_string(candidate.upper())
        except ValueError:
            return None

    def add_column(self):
        worksheet = self._get_sheet()
        if worksheet is None:
            messagebox.showwarning("No Sheet", "Select a workbook and sheet first.")
            return
        column_name = simpledialog.askstring("Add Column", "Column name (optional):")
        index = max(worksheet.max_column, 1) + 1
        if column_name:
            header_cell = worksheet.cell(row=1, column=index)
            if header_cell.value not in (None, "") and not messagebox.askyesno(
                "Overwrite Header",
                "The first row already has data for this column. Overwrite it with the new column name?",
            ):
                return
            header_cell.value = column_name.strip()
        self._render_sheet()

    def delete_column(self):
        worksheet = self._get_sheet()
        if worksheet is None:
            messagebox.showwarning("No Sheet", "Select a workbook and sheet first.")
            return
        target = simpledialog.askstring("Delete Column", "Enter column letter or index (e.g., A or 1):")
        if not target:
            return
        index = self._parse_column_index(target)
        if index is None or index > max(worksheet.max_column, 1):
            messagebox.showwarning("Invalid Column", "Provide a valid existing column letter or index.")
            return
        if not messagebox.askyesno("Confirm Delete", f"Delete column {get_column_letter(index)}?"):
            return
        worksheet.delete_cols(index, 1)
        self._render_sheet()

    def edit_cell(self, event):
        worksheet = self._get_sheet()
        if worksheet is None:
            return
        item_id = self.tree.identify_row(event.y)
        column_id = self.tree.identify_column(event.x)
        if not item_id or not column_id:
            return

        view_row_index = self.tree.index(item_id)
        if view_row_index >= len(self.row_index_map):
            return
        excel_row = self.row_index_map[view_row_index]
        excel_col = int(column_id.replace("#", ""))
        current_value = worksheet.cell(row=excel_row, column=excel_col).value
        new_value = simpledialog.askstring(
            "Edit Cell",
            f"Update {get_column_letter(excel_col)}{excel_row}",
            initialvalue="" if current_value is None else str(current_value),
        )
        if new_value is None:
            return
        worksheet.cell(row=excel_row, column=excel_col).value = new_value
        self._render_sheet()

    def refresh(self):
        self._refresh_file_list()
        self._render_sheet()
        if self.on_change:
            self.on_change()

    def _on_file_select(self, _event=None):
        selected = self.file_listbox.curselection()
        if not selected:
            return
        path = self.file_paths[selected[0]]
        self._set_current_file(path)

    def _on_sheet_select(self, _event=None):
        workbook = self._get_workbook()
        if not workbook:
            return
        chosen = self.sheet_var.get()
        if chosen in workbook.sheetnames:
            self.current_sheet = chosen
            self._render_sheet()

    def get_loaded_file_count(self):
        return len(self.workbooks)
